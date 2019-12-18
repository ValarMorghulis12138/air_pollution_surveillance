import os
try:
    from osgeo import gdal
    from osgeo import ogr
    from osgeo import osr
except ImportError:
    import gdal
    import ogr
    import osr
# import osgeo.ogr as ogr
# import osgeo.osr as osr
import openpyxl


def main(input_xlsx, out_raster, china_shp):
    # input_xlsx = 'pm25_geoed.xlsx'
    # use openpyxl so we can access by field name
    wb = openpyxl.load_workbook(input_xlsx)
    ws = wb.get_sheet_by_name('pm25')
    max_row = ws.max_row

    gdal.AllRegister()
    # 为了支持中文路径，请添加下面这句代码
    gdal.SetConfigOption("GDAL_FILENAME_IS_UTF8", "NO")

    # 为了使属性表字段支持中文，请添加下面这句
    gdal.SetConfigOption("SHAPE_ENCODING", "CP936")

    # 注册所有的驱动
    ogr.RegisterAll()
    
    # set up the shapefile driver
    driver = ogr.GetDriverByName("ESRI Shapefile")

    # create the data source
    output_shp = './geoed_excel_to_raster/air_pollution.shp'
    data_source = driver.CreateDataSource(output_shp)

    # create the spatial reference, WGS84
    srs = osr.SpatialReference()
    srs.ImportFromEPSG(4326)

    # create the layer
    layer = data_source.CreateLayer("surveillance_stations", srs, ogr.wkbPoint)

    # Add the fields we're interested in
    field_name = ogr.FieldDefn("Name", ogr.OFTString)
    field_name.SetWidth(24)
    layer.CreateField(field_name)

    field_city = ogr.FieldDefn("City", ogr.OFTString)
    field_city.SetWidth(24)
    layer.CreateField(field_city)

    field_airquality = ogr.FieldDefn("AQ", ogr.OFTString)
    field_airquality.SetWidth(24)
    layer.CreateField(field_airquality)

    field_primepollution = ogr.FieldDefn("PP", ogr.OFTString)
    field_primepollution.SetWidth(24)
    layer.CreateField(field_primepollution)

    field_time = ogr.FieldDefn("Time", ogr.OFTString)
    field_time.SetWidth(24)
    layer.CreateField(field_time)

    layer.CreateField(ogr.FieldDefn("Latitude", ogr.OFTReal))
    layer.CreateField(ogr.FieldDefn("Longitude", ogr.OFTReal))
    layer.CreateField(ogr.FieldDefn("Pm2.5", ogr.OFTInteger))
    layer.CreateField(ogr.FieldDefn("Pm10", ogr.OFTInteger))
    layer.CreateField(ogr.FieldDefn("CO", ogr.OFTReal))
    layer.CreateField(ogr.FieldDefn("NO2", ogr.OFTInteger))
    layer.CreateField(ogr.FieldDefn("O3/1HOUR", ogr.OFTInteger))
    layer.CreateField(ogr.FieldDefn("O3/8HOUR", ogr.OFTInteger))
    layer.CreateField(ogr.FieldDefn("SO2", ogr.OFTInteger))

    # Process the xlsx file and add the attributes and features to the shpfile
    for i in range(2, max_row+1):
        lon = ws['O' + str(i)].value
        lat = ws['P' + str(i)].value

        if (lon is not None) and (lat is not None):
            # create the feature
            feature = ogr.Feature(layer.GetLayerDefn())
            # Set the attributes using the values from the delimited text file
            feature.SetField("Name", ws['C' + str(i)].value)
            feature.SetField("City", ws['B' + str(i)].value)
            feature.SetField("AQ", ws['E' + str(i)].value)
            feature.SetField("PP", ws['F' + str(i)].value)
            feature.SetField("Time", ws['N' + str(i)].value)
            feature.SetField("Latitude", ws['P' + str(i)].value)
            feature.SetField("Longitude", ws['O' + str(i)].value)
            feature.SetField("Pm2.5", ws['G' + str(i)].value)
            feature.SetField("Pm10", ws['H' + str(i)].value)
            feature.SetField("CO", ws['I' + str(i)].value)
            feature.SetField("NO2", ws['J' + str(i)].value)
            feature.SetField("O3/1HOUR", ws['K' + str(i)].value)
            feature.SetField("O3/8HOUR", ws['L' + str(i)].value)
            feature.SetField("SO2", ws['M' + str(i)].value)
            # create the WKT for the feature using Python string formatting
            wkt = "POINT(%f %f)" % (float(lon), float(lat))

            # Create the point from the Well Known Txt
            point = ogr.CreateGeometryFromWkt(wkt)

            # Set the feature geometry using the point
            feature.SetGeometry(point)
            # Create the feature in the layer (shapefile)
            layer.CreateFeature(feature)
            # Dereference the feature
            feature = None

    # Save and close the data source
    data_source = None

    # convert shapefile to raster using gdal
    raw_raster = './geoed_excel_to_raster/raw_ratser.tif'
    gdal.Grid(raw_raster, output_shp, format='GTiff', zfield='pm2.5',
              outputBounds=[73.4, 54, 135.1, 18.1],
              algorithm='invdist:power=2.0:smoothing=0.1')
    # clip to china extent using gdal
    gdal.Warp(out_raster, raw_raster, format="GTiff",
              warpOptions="NUM_THREADS=ALL_CPUS", warpMemoryLimit=1024,
              resampleAlg="cubic", multithread=True,
              cutlineDSName=china_shp, cropToCutline=True)
    os.remove(raw_raster)


if __name__ == '__main__':
    input_xlsx = './geoed_excel_to_raster/pm25_geoed.xlsx'
    output_raster = './geoed_excel_to_raster/air_pollution_pm2p5.tif'
    china_shp = './geoed_excel_to_raster/china_extent.shp'
    main(input_xlsx, output_raster, china_shp)