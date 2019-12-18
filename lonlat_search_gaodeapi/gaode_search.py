import requests
import openpyxl


def getGeoForAddress(address):
    # address = "上海市中山北一路121号"
    params = {
        'key': '913a3a4b40cc9e2a8757a899a337aff6',
        'address': address
    }
    ret = requests.get("https://restapi.amap.com/v3/geocode/geo?", params=params)
    # print(ret.url)        # 打印访问的url    http://httpbin.org/get?key1=value1&key2=value2
    # print(ret.text)       # 打印返回值
    # print(ret.json())     # 将返回的JSON数据
    # print(ret.content)    # 打印二进制的数据格式，转化为str需要decode转码
    contest = ret.json()
    try:
        geo = contest['geocodes'][0]['location']
        # print(type(geo))        # geo type class str
        lon = geo.split(',')[0]
        lat = geo.split(',')[1]
    except Exception as e:
        print(f"Unexpected error: {e}")
        lon = None
        lat = None
    finally:
        return lon, lat


def getGeoForAddressCity(address, city):
    params = {
        'key': '913a3a4b40cc9e2a8757a899a337aff6',
        'address': address,
        'city': city
    }
    ret = requests.get("https://restapi.amap.com/v3/geocode/geo?", params=params)
    # print(ret.url)        # 打印访问的url    http://httpbin.org/get?key1=value1&key2=value2
    # print(ret.text)       # 打印返回值
    # print(ret.json())     # 将返回的JSON数据
    # print(ret.content)    # 打印二进制的数据格式，转化为str需要decode转码
    contest = ret.json()
    try:
        geo = contest['geocodes'][0]['location']
        # print(type(geo))        # geo type class str
        lon = geo.split(',')[0]
        lat = geo.split(',')[1]
    except Exception as e:
        print(f"Unexpected error: {e}")
        lon = None
        lat = None
    finally:
        return lon, lat


if __name__ == '__main__':
    input_xlsx = 'pm25.xlsx'
    output_xlsx = 'pm25_geoed.xlsx'

    wb = openpyxl.load_workbook(input_xlsx)
    ws = wb.get_sheet_by_name('pm25')
    max_row = ws.max_row
    ws['O1'].value = 'Lon'
    ws['P1'].value = 'Lat'
    for i in range(2, max_row+1):
        city = ws['B' + str(i)].value
        address = ws['C' + str(i)].value
        lon, lat = getGeoForAddressCity(address, city)
        if (lon is not None) and (lat is not None):
            ws['O' + str(i)].value = lon
            ws['P' + str(i)].value = lat

    wb.save(output_xlsx)
